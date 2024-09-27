import openpyxl
import openpyxl.cell
import openpyxl.workbook
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import argparse
import os
import sys

################################################################################### CLASSES ###################################################################################

class BColors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

class ColoredMessage:
  def success(message):
    return f'{BColors.OKGREEN}{message}{BColors.ENDC}'

  def processing(message):
    return f'{BColors.OKCYAN}{message}{BColors.ENDC}'

  def error(message):
    return f'{BColors.FAIL}{message}{BColors.ENDC}'

  def warning(message):
    return f'{BColors.WARNING}{message}{BColors.ENDC}'

class ProgressBar:
  def __init__(self, max_i):
    self.max_i = max_i

  def progress(self, i):
    progress = int(100*i/self.max_i)
    remaining = 100 - progress
    self.progress_bar = f"[{ColoredMessage.success(progress*'#')}{ColoredMessage.processing(remaining*'-')}]"

  def print(self):
    print(self.progress_bar, end='\r', sep='')

class SheetWithAttributes:
  def __init__(self, filename: str, sheet: openpyxl.worksheet.worksheet.Worksheet, indexed_identifiers: dict[int, str], min_row: int = None, max_row: int = None, min_col: int = None, max_column: int = None):
    self.filename = filename
    self.sheet = sheet
    self.min_row = min_row or 2
    self.max_row = max_row or sheet.max_row
    self.min_col = min_col or 1
    self.max_column = max_column or sheet.max_column
    self.identifiers = indexed_identifiers
    self.col_indexes_per_title = self.get_col_indexes_per_title()
    self.identifiers_indexes = self.get_identifiers_indexes()

  def get_rows_generator(self):
    return self.sheet.iter_rows(min_row=self.min_row, max_row=self.max_row, min_col=self.min_col, max_col=self.max_column)

  def get_headers_generator(self):
      return self.sheet.iter_cols(min_row=1, max_row=1, min_col=self.min_col, max_col=self.max_column)

  def get_col_indexes_per_title(self):
    sheet_cols_generator = self.get_headers_generator()
    index_range = range(0, self.max_column - 1)
    dict = {}
    for col, i in zip(sheet_cols_generator, index_range):
      if col[0].value in dict:
        print(ColoredMessage.warning(f'Column \'{col[0].value}\' duplicated in \'{self.filename}\', using the first one'))
        continue
      dict[col[0].value] = i
    return dict

  def get_identifiers_indexes(self):
    dict = {}
    for identifier in self.identifiers.values():
      if identifier in self.col_indexes_per_title:
        dict[identifier] = self.col_indexes_per_title[identifier]
      else:
        raise NameError(ColoredMessage.error(f'Column \'{identifier}\' not found in \'{self.filename}\'')) from None
    return dict

################################################################################ FUNCTIONS ################################################################################

def get_sheet_with_attributes_from_workbook(filename, wb: openpyxl.workbook.workbook.Workbook, sheet_index: int, identifiers: dict[int, str], min_row: int = None, max_row: int = None, min_col: int = None, max_col: int = None):
  sheetname = wb.sheetnames[sheet_index]
  sheet = wb[sheetname]
  return SheetWithAttributes(filename, sheet, identifiers, min_row, max_row, min_col, max_col)

def is_corrispondent_row(sheet_1: SheetWithAttributes, sheet_1_row: dict[str, any], sheet_2: SheetWithAttributes, sheet_2_row: dict[str, any]):
  for i, source_identifier in sheet_1.identifiers.items():
    dest_identifier = sheet_2.identifiers[i]
    if sheet_2.identifiers[i] not in sheet_2_row:
      raise NameError(ColoredMessage.error(f'Column to match \'{dest_identifier}\' not found in dest sheet')) from None
    if sheet_1_row[source_identifier] != sheet_2_row[dest_identifier]:
      return False
  return True

def find_corrispondent_rows(source: SheetWithAttributes, source_rows: list[dict[str, any]], dest: SheetWithAttributes, dest_rows: list[dict[str, any]]):
  rows_corrispondence = []
  progress_bar = ProgressBar(len(source_rows))
  for source_i, source_row in enumerate(source_rows):
    for dest_i, dest_row in enumerate(dest_rows):
      if is_corrispondent_row(source, source_row, dest, dest_row):
        rows_corrispondence.append((source_i, dest_i))
    progress_bar.progress(source_i)
    progress_bar.print()
  return rows_corrispondence

def copy_col_to_dest(source: SheetWithAttributes, dest: SheetWithAttributes, row_corrispondence: tuple[int, int], to_copy_identifier: str, to_paste_identifier: str):
  source_row_i, dest_row_i = row_corrispondence
  if to_copy_identifier not in source.col_indexes_per_title:
    raise NameError(ColoredMessage.error(f'Column \'{to_copy_identifier}\' to write not found in \'{source.filename}\'')) from None
  to_copy_index = source.col_indexes_per_title[to_copy_identifier]
  if to_paste_identifier not in dest.col_indexes_per_title:
    raise NameError(ColoredMessage.error(f'Column \'{to_paste_identifier}\' to write not found in \'{dest.filename}\'')) from None
  to_paste_index = dest.col_indexes_per_title[to_paste_identifier]
  source_cell = source.sheet.cell(row=source_row_i+2, column=to_copy_index+1)
  dest_cell = dest.sheet.cell(row=dest_row_i+2, column=to_paste_index+1)
  dest_cell.value = source_cell.value

def copy_cols_to_dest(source: SheetWithAttributes, dest: SheetWithAttributes, row_corrispondence: tuple[int, int], source_to_copy_identifiers: dict[int, str], dest_to_paste_identifiers: dict[int, str]):
  for i, to_copy_identifier in source_to_copy_identifiers.items():
    to_paste_identifier = dest_to_paste_identifiers[i]
    copy_col_to_dest(source, dest, row_corrispondence, to_copy_identifier, to_paste_identifier)

def sheet_to_list_of_row_dicts(sheet: SheetWithAttributes):
  rows_generator = sheet.sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)
  return [{identifier: row[i].value for identifier, i in sheet.identifiers_indexes.items()} for row in rows_generator]

def write_results_to_sheet(source, dest, rows_corrispondences, source_to_copy_identifiers, dest_to_paste_identifiers):
  progress_bar = ProgressBar(len(rows_corrispondences))
  for i, row_corrispondence in enumerate(rows_corrispondences):
    copy_cols_to_dest(source, dest, row_corrispondence, source_to_copy_identifiers, dest_to_paste_identifiers)
    progress_bar.progress(i)
    progress_bar.print()

def loadWbOrRaise(filename):
  try:
    return openpyxl.load_workbook(filename)
  except FileNotFoundError:
    raise FileNotFoundError(ColoredMessage.error(f'File \'{filename}\' not found, check if the filename is correct and the file is placed in the script root folder')) from None

def save_results(dest_wb, output_dir, output_filename):
  if not os.path.isdir(output_dir):
    os.mkdir(output_dir)
  dest_wb.save(os.path.join(output_dir, output_filename))

def run_process(args):
  source_filename = args.source_filename
  dest_filename = args.dest_filename
  output_dir = args.output_dir
  output_filename = args.output_filename
  source_offset = args.source_offset
  source_limit = args.source_limit
  source_max_row = (source_offset + source_limit) if source_limit else None

  source_to_match_identifiers = { i: identifier for i, identifier in enumerate(args.source_to_match_identifiers) }
  dest_to_match_identifiers = { i: identifier for i, identifier in enumerate(args.dest_to_match_identifiers) }
  source_to_copy_identifiers = { i: identifier for i, identifier in enumerate(args.source_to_copy_identifiers) }
  dest_to_paste_identifiers = { i: identifier for i, identifier in enumerate(args.dest_to_paste_identifiers) }

  if len(source_to_match_identifiers) != len(dest_to_match_identifiers):
    raise NameError(ColoredMessage.error('Different number of column to match, they must have same number of elements and in order of correlation')) from None
  if len(source_to_copy_identifiers) != len(dest_to_paste_identifiers):
    raise NameError(ColoredMessage.error('Different number of column for copy, they must have same number of elements and in order of correlation')) from None

  source_wb = loadWbOrRaise(source_filename)
  dest_wb = loadWbOrRaise(dest_filename)

  source = get_sheet_with_attributes_from_workbook(source_filename, source_wb, 0, source_to_match_identifiers, source_offset, source_max_row)
  dest = get_sheet_with_attributes_from_workbook(dest_filename,dest_wb, 0, dest_to_match_identifiers)

  source_rows = sheet_to_list_of_row_dicts(source)
  dest_rows = sheet_to_list_of_row_dicts(dest)

  print(ColoredMessage.processing('Reading files...\n'))
  rows_corrispondences = find_corrispondent_rows(source, source_rows, dest, dest_rows)
  print(ColoredMessage.success('\n\nRead completed\n'))

  print(ColoredMessage.processing('Writing result...\n'))
  write_results_to_sheet(source, dest, rows_corrispondences, source_to_copy_identifiers, dest_to_paste_identifiers)

  save_results(dest_wb, output_dir, output_filename)
  print(ColoredMessage.success(f'\n\nReport successfully fullfilled, saved to {output_filename}'))

if __name__ == '__main__':
  parser = argparse.ArgumentParser(description='Copy selected columns data from one sheet to another, based on equal identifiers')
  parser.add_argument('-s', '--source-filename', default='source.xlsx', type=str, help='Source xlsx file where to find values to copy | Default = \'source.xlsx\'')
  parser.add_argument('-d', '--dest-filename', default='dest.xlsx', type=str, help='Dest xlsx file where to copy found values | Default = \'dest.xlsx\'')
  parser.add_argument('-o', '--output-filename', default=f'output.xlsx', type=str, help='Output xlsx file to produce with results | Default = \'output.xlsx\'')
  parser.add_argument('-od', '--output-dir', default='Output', type=str, help='Output folder where to place output file | Default = \'Output\'')
  parser.add_argument('-si', '--source-to-match-identifiers', nargs='+', required=True, help='Columns titles of the values to match in the source sheet')
  parser.add_argument('-di', '--dest-to-match-identifiers', nargs='+', required=True, help='Columns titles of the values to match in the dest sheet, in the same order of the relative source identifiers')
  parser.add_argument('-sc', '--source-to-copy-identifiers', nargs='+', required=True, help='Columns titles of the values of the values to copy from the source sheet')
  parser.add_argument('-sp', '--dest-to-paste-identifiers', nargs='+', required=True, help='Columns titles of the values of the values to paste in the dest sheet, in the same order of the relative source identifiers')
  parser.add_argument('-so', '--source-offset', default=0, type=int, help='Number of row to ignore from start  | Default = 0')
  parser.add_argument('-sl', '--source-limit', default=None, type=int, help='Number of row to process | Default = None (continue until last sheet row)')
  parser.add_argument('-deb', '--debug', action='store_true', help='Print debug messages')

  args = parser.parse_args()

  if not args.debug:
    sys.tracebacklimit = 0

  run_process(args)
