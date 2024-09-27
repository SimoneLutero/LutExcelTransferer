from openpyxl.worksheet.worksheet import Worksheet
from colored_message import ColoredMessage

class SheetWithAttributes:
  def __init__(self, filename: str, sheet: Worksheet, indexed_identifiers: dict[int, str], min_row: int = None, max_row: int = None, min_col: int = None, max_column: int = None):
    self.filename = filename
    self.sheet = sheet
    self.min_row = min_row or 2
    self.max_row = max_row or sheet.max_row
    self.min_col = min_col or 1
    self.max_column = max_column or sheet.max_column
    self.identifiers = indexed_identifiers
    self.col_indexes_per_title = self.__get_col_indexes_per_title()
    self.identifiers_indexes = self.__get_identifiers_indexes()

  def get_rows_generator(self):
    return self.sheet.iter_rows(min_row=self.min_row, max_row=self.max_row, min_col=self.min_col, max_col=self.max_column)

  def __get_headers_generator(self):
    return self.sheet.iter_cols(min_row=1, max_row=1, min_col=self.min_col, max_col=self.max_column)

  def __get_col_indexes_per_title(self):
    sheet_cols_generator = self.__get_headers_generator()
    index_range = range(0, self.max_column - 1)
    dict = {}
    for col, i in zip(sheet_cols_generator, index_range):
      if col[0].value not in dict:
        dict[col[0].value] = i
      else:
        print(ColoredMessage.warning(f'Column \'{col[0].value}\' duplicated in \'{self.filename}\', using the first one'))
    return dict

  def __get_identifiers_indexes(self):
    dict = {}
    for identifier in self.identifiers.values():
      if identifier in self.col_indexes_per_title:
        dict[identifier] = self.col_indexes_per_title[identifier]
      else:
        raise NameError(ColoredMessage.error(f'Column \'{identifier}\' not found in \'{self.filename}\'')) from None
    return dict